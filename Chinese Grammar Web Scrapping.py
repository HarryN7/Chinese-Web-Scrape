from lxml import html
import requests
import openpyxl
from openpyxl.styles import Font

#Open up new workbook and select default workbook
wb = openpyxl.Workbook()

def text(elt):  #Function to convert html elements to text
    return elt.text_content().replace(u'\xa0', u' ') #Replace unicode spaces with real spaces

def formatHeader(ws):   #Function to format headers in each worksheet
    headerRange = ['A1', 'B1', 'C1', 'D1', 'E1']
    for pos in headerRange:
        ws[pos].font = Font(size=12, bold=True)
    ws.auto_filter.ref = ws.dimensions  #adds filter to table in worksheet

def webCrawl(rootURL, links):   #Crawl through all links from source URL and grab first paragraph
    print("Beginning Web Crawl...")
    pages = [requests.get(rootURL + link) for link in links]    #Grab each page from each link
    trees = [html.fromstring(page.content) for page in pages]   #Grab each tree from each page
    paragraphs = [text(tree.xpath('//p[1]')[0]) for tree in trees]  #Grab the first paragraph from each tree
    print("Web Crawl Complete!")
    return paragraphs

def extractGrammarPoints(sourceURL, wsTitle):   #Extract grammar points from URL and write to excel
    page = requests.get(sourceURL)
    tree = html.fromstring(page.content)
    rootURL = "https://resources.allsetlearning.com"
    headers = ["Category"]
    tableNo = 0
    i = 0
    ws = wb['Sheet']

    categories = [text(category) for category in tree.xpath('//html/body//*[self::h3 or self::h4]')]    #Get all h3 and h4 nodes
    links = [str(link) for link in tree.xpath('//table[@class="wikitable"]//a/@href')]  #Get all links from sourceURL and convert to string
    rowsHeaders = [text(header) for header in tree.xpath('//table[@class="wikitable"][1]//th')] #Get headers from first table
    headers.extend(rowsHeaders) #Add row headers as a list to the headers list
    headers.extend(["Description"])
    ws.append(headers)
    paragraphs = webCrawl(rootURL, links)

    #Iterate through all table rows in relevant tables and extract data
    for tr in tree.xpath('//table[@class="wikitable"]//tr'):
        if "Grammar Point" in text(tr):
            tableNo+=1
        else:
            rowValues = [text(td) for td in tr.xpath('td')]
            if tableNo <= len(categories):
                data = [categories[tableNo - 1]]
            data.append('=HYPERLINK("' + rootURL + links[i] + '","' + rowValues[0].replace('"','\'') + '")')
            data.extend(rowValues[1:3]) #Add table data into data list
            data.append(paragraphs[i])
            i += 1
            ws.append(data)
    
    formatHeader(ws)
    ws.title = wsTitle

extractGrammarPoints("https://resources.allsetlearning.com/chinese/grammar/A2_grammar_points", "A2 Grammar Points")
wb.create_sheet('Sheet')
extractGrammarPoints("https://resources.allsetlearning.com/chinese/grammar/B1_grammar_points", "B1 Grammar Points")
wb.create_sheet('Sheet')
extractGrammarPoints("https://resources.allsetlearning.com/chinese/grammar/B2_grammar_points", "B2 Grammar Points")

wb.save('Chinese Grammar Points.xlsx')